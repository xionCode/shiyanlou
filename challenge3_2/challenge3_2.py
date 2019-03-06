# -*- coding: utf-8 -*-

from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
import datetime # 可以用来处理时间相关的数据

wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    '''
    该函数可以用来处理原数据文件：
    1. 合并表格并写入的 combine 表中
    2. 保存原数据文件
    '''
    combine_sheet = wb.copy_worksheet(students_sheet)
    combine_sheet.title = 'combine'
    combine_sheet['D1'] = '学习时间'
    for row in combine_sheet.iter_rows(2):
        course_name = row[1].value
        cell_row = row[1].row

        for row in time_sheet.iter_rows(2):
            if row[1].value == course_name:
                combine_sheet.cell(row=cell_row, column=4,value=row[2].value)
                break
    
    wb.save('courses.xlsx')

def split():
    '''
    该函数可以用来分割文件：
    1. 读取 combine 表中的数据
    2. 将数据按时间分割
    3. 写入不同的数据表中
    '''
    combine_sheet = wb['combine']
    year_list = []
    for row in combine_sheet.values:
        if row[0] != '创建时间':
            year_list.append(row[0].strftime('%Y'))
    
    for year in set(year_list):
        wb_temp = Workbook()
        ws = wb_temp.active
        ws.title = year
        ws.append(['创建时间','课程名称','学习人数','学习时间'])
        for row in combine_sheet.values:
            if row[0] != '创建时间':
                if row[0].strftime('%Y') == year:
                    ws.append(row)
        
        wb_temp.save('{}.xlsx'.format(year))
    

# 执行
if __name__ == '__main__':
    combine()
    split()